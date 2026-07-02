# -*- coding: utf-8 -*-
from __future__ import annotations
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from utils import normalize_excel_key
from sent_parser import parse_sent_html_file, parse_sent_xml_file
import re

HEADER_ROW = 7

def excel_path_for_client(base_dir: Path, client) -> Path:
    p = Path(client.excel_file)
    if not p.is_absolute():
        p = base_dir / "excels" / p
        if not p.exists():
            p = base_dir / client.excel_file
    return p

def sent_folder_for_client(base_dir: Path, client) -> Path:
    return base_dir / client.sent_folder

def load_excel(base_dir: Path, client) -> pd.DataFrame:
    path = excel_path_for_client(base_dir, client)
    if not path.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku Excel: {path}")
    df = pd.read_excel(path, header=HEADER_ROW - 1)
    df.columns = [str(c).strip() for c in df.columns]
    return df

def find_row_by_lp(df: pd.DataFrame, lp: str):
    if "LP" not in df.columns:
        raise ValueError("Brak kolumny 'LP' w arkuszu.")
    tmp = df.copy()
    tmp["LP_str"] = tmp["LP"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    result = tmp.loc[tmp["LP_str"] == str(lp).strip()]
    if result.empty:
        raise ValueError(f"Nie znaleziono LP={lp} w kolumnie 'LP'.")
    return result.iloc[0]

def detect_sent_type(row, client) -> str:
    rule = client.sent_type_rule
    if not rule:
        return client.default_sent_type
    value = ""
    if rule.column_name and rule.column_name in row.index:
        value = str(row.get(rule.column_name, ""))
    elif rule.column_index is not None and len(row) > rule.column_index:
        value = str(row.iloc[rule.column_index])
    return rule.true_type if rule.contains.upper() in value.upper() else rule.false_type

def get_headers_from_excel(path: Path) -> list[str]:
    df = pd.read_excel(path, header=HEADER_ROW - 1, nrows=0)
    return [str(c).strip() for c in df.columns]

def extract_sent_from_filename(file_name) -> str:
    match = re.search(r"(SENT\d+)", str(file_name or ""), re.IGNORECASE)
    return match.group(1).upper() if match else ""


COUNTRY_MAP = {
    "POLSKA": "PL", "POLAND": "PL",
    "NIEMCY": "DE", "GERMANY": "DE", "DEUTSCHLAND": "DE",
    "NORWEGIA": "NO", "NORWAY": "NO", "NORGE": "NO",
    "SZWECJA": "SE", "SWEDEN": "SE", "SVERIGE": "SE",
    "DANIA": "DK", "DENMARK": "DK",
    "ISLANDIA": "IS", "ICELAND": "IS",
    "WIELKA BRYTANIA": "GB", "UK": "GB", "UNITED KINGDOM": "GB",
}

def _cell_text(ws, address: str) -> str:
    try:
        value = ws[str(address or "").strip() or "A1"].value
    except Exception:
        value = None
    return " ".join(str(value or "").replace("\xa0", " ").split()).strip()

def _clean_sender_name(value: str) -> str:
    text = " ".join(str(value or "").split()).strip(" -:")
    text = re.sub(r"^AWIZACJE\s+", "", text, flags=re.IGNORECASE).strip(" -:")
    return text

def _parse_identity(value: str) -> tuple[str, str, str]:
    text = " ".join(str(value or "").split()).strip()
    if not text:
        return "INNY", "", ""
    producer_name = ""
    if "PRODUCENT" in text.upper() or "NADAWCA" in text.upper():
        # np. Producent: (...) Example Producer Ltd. DE000000000
        after = re.split(r"\)", text, maxsplit=1)[-1].strip() if ")" in text else text.split(":", 1)[-1].strip()
        m = re.search(r"\b([A-Z]{2}\s?\d[A-Z0-9 .\-/]{4,})$", after.strip(), re.IGNORECASE)
        if m:
            identity = m.group(1).strip()
            producer_name = after[:m.start()].strip(" ,:-")
            id_type = "VAT UE" if identity.upper().startswith(("DE", "DK", "SE")) else "INNY"
            return id_type, identity, producer_name
    if ":" in text:
        left, right = text.split(":", 1)
        left_u = left.upper()
        if "VAT" in left_u and "UE" in left_u:
            id_type = "VAT UE"
        elif "NIP" in left_u:
            id_type = "NIP"
        else:
            id_type = "INNY"
        return id_type, right.strip(), producer_name
    # fallback: ostatni token wygladajacy jak identyfikator
    m = re.search(r"\b([A-Z]{0,2}\s?[0-9][A-Z0-9 .\-/]{4,})\b", text, re.IGNORECASE)
    return "INNY", (m.group(1).strip() if m else text), producer_name

def _parse_address(value: str) -> dict[str, str]:
    text = " ".join(str(value or "").split()).strip()
    result = {"street": "", "house_number": "", "postal_code": "", "city": "", "country": ""}
    if not text:
        return result
    parts = [p.strip() for p in text.split(",") if p.strip()]
    # Jezeli pierwszy fragment wyglada jak nazwa firmy bez numeru domu, pomin go.
    if len(parts) >= 2 and not re.search(r"\d", parts[0]):
        parts = parts[1:]
    street_part = parts[0] if parts else text
    m = re.match(r"^(.*?)(?:\s+|^)(\d+[A-Za-z]?(?:[/-]\d+[A-Za-z]?)?)$", street_part)
    if m:
        result["street"] = m.group(1).strip(" ,")
        result["house_number"] = m.group(2).strip()
    else:
        result["street"] = street_part.strip(" ,")
    rest = " ".join(parts[1:]) if len(parts) > 1 else ""
    m2 = re.search(r"\b([A-Z]{1,3}-?\d{2,6}|\d{2}-\d{3}|\d{4,6})\b\s*(.*)", rest, re.IGNORECASE)
    if m2:
        postal = m2.group(1).strip()
        result["postal_code"] = postal
        city_country = m2.group(2).strip(" ,")
        # Kraj moze byc ostatnim slowem.
        bits = city_country.split()
        if bits:
            last = bits[-1].upper().strip(".,")
            if last in COUNTRY_MAP or re.fullmatch(r"[A-Z]{2}", last):
                result["country"] = COUNTRY_MAP.get(last, last)
                city_country = " ".join(bits[:-1]).strip(" ,")
        result["city"] = city_country
        if not result["country"] and "-" in postal and re.match(r"^[A-Z]{1,3}-", postal):
            prefix = postal.split("-", 1)[0].upper()
            if len(prefix) == 1:
                prefix = {"D": "DE"}.get(prefix, prefix)
            result["country"] = prefix
    return result

def read_sender_from_excel(base_dir: Path, client) -> dict[str, str]:
    """Czyta dane GoodsSender z komorek wskazanych w konfiguracji klienta, domyslnie A1/A2/A3."""
    if getattr(client, "sender_source", "mapping") != "excel_top_cells":
        return {}
    path = excel_path_for_client(base_dir, client)
    if not path.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku Excel: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw_name = _cell_text(ws, getattr(client, "sender_name_cell", "A1"))
    raw_identity = _cell_text(ws, getattr(client, "sender_identity_cell", "A2"))
    raw_address = _cell_text(ws, getattr(client, "sender_address_cell", "A3"))
    identity_type, identity_number, producer_name = _parse_identity(raw_identity)
    address = _parse_address(raw_address)
    name = producer_name or _clean_sender_name(raw_name)
    if not address.get("country") and identity_number:
        prefix_match = re.match(r"^([A-Z]{2})", identity_number.strip().upper())
        if prefix_match:
            address["country"] = prefix_match.group(1)

    return {
        "GoodsSender/TraderInfo/TraderName": name,
        "GoodsSender/TraderInfo/TraderIdentityType": identity_type or "INNY",
        "GoodsSender/TraderInfo/TraderIdentityNumber": identity_number,
        "GoodsSender/TraderAddress/Street": address.get("street", ""),
        "GoodsSender/TraderAddress/HouseNumber": address.get("house_number", ""),
        "GoodsSender/TraderAddress/City": address.get("city", ""),
        "GoodsSender/TraderAddress/Country": address.get("country", ""),
        "GoodsSender/TraderAddress/PostalCode": address.get("postal_code", ""),
    }

def import_sent_data(base_dir: Path, client, log=lambda m: None) -> str:
    excel_path = excel_path_for_client(base_dir, client)
    sent_folder = sent_folder_for_client(base_dir, client)
    if not excel_path.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku Excel: {excel_path.name}")
    if not sent_folder.exists():
        raise FileNotFoundError(f"Nie znaleziono folderu: {sent_folder}")
    files = sorted(list(sent_folder.glob("*.html")) + list(sent_folder.glob("*.htm")) + list(sent_folder.glob("*.xml")))
    if not files:
        raise FileNotFoundError(f"Folder {sent_folder.name} nie zawiera plików HTML ani XML.")
    wb = load_workbook(excel_path)
    ws = wb.active
    headers = {normalize_excel_key(c.value): c.column for c in ws[HEADER_ROW] if normalize_excel_key(c.value)}
    required = ["LP", "SENT", "KLUCZ ODBIORCY", "NR CMR PRZYPISANY DO SENT"]
    missing = [x for x in required if x not in headers]
    if missing:
        raise ValueError("Brak wymaganych kolumn w Excelu: " + ", ".join(missing))
    weight_col = headers.get("WAGA ATS") or headers.get("WAGA HKP")
    row_by_lp = {}
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        lp = normalize_excel_key(ws.cell(row_idx, headers["LP"]).value)
        if lp:
            row_by_lp[lp] = row_idx
    updated, skipped = 0, []
    for path in files:
        name_key = path.name.upper().replace("_", "").replace(" ", "")
        if "SENT215" not in name_key:
            skipped.append(f"{path.name}: pominięto — nie jest SENT215")
            continue
        data = parse_sent_xml_file(path) if path.suffix.lower() == ".xml" else parse_sent_html_file(path)
        lp = normalize_excel_key(data.get("lp"))
        if not lp:
            skipped.append(f"{path.name}: brak Numeru własnego zgłoszenia")
            continue
        if lp not in row_by_lp:
            skipped.append(f"{path.name}: nie znaleziono LP={lp} w Excelu")
            continue
        r = row_by_lp[lp]
        if data.get("sent"):
            ws.cell(r, headers["SENT"]).value = data["sent"]
        if data.get("recipient_key"):
            ws.cell(r, headers["KLUCZ ODBIORCY"]).value = data["recipient_key"]
        if data.get("cmr"):
            ws.cell(r, headers["NR CMR PRZYPISANY DO SENT"]).value = data["cmr"]
        if weight_col and data.get("weight") != "":
            ws.cell(r, weight_col).value = data["weight"]
        updated += 1
        log(f"OK {path.name}: LP={lp}, SENT={data.get('sent')}, CMR={data.get('cmr')}")
    wb.save(excel_path)
    msg = f"Zaktualizowano {updated} wierszy w pliku {excel_path.name}."
    if skipped:
        msg += "\n\nPominięte pliki:\n" + "\n".join(skipped[:30])
    return msg

def update_close_dates(base_dir: Path, client, log=lambda m: None) -> str:
    excel_path = excel_path_for_client(base_dir, client)
    folder = sent_folder_for_client(base_dir, client)
    if not excel_path.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku Excel: {excel_path.name}")
    if not folder.exists():
        raise FileNotFoundError(f"Nie znaleziono folderu: {folder}")
    wb = load_workbook(excel_path)
    ws = wb.active
    headers = {normalize_excel_key(c.value): c.column for c in ws[HEADER_ROW] if normalize_excel_key(c.value)}
    for key in ["SENT", "DATA ZAMKNIĘCIA SENT"]:
        if key not in headers:
            raise ValueError(f"Brak kolumny: {key}")
    row_by_sent = {}
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        sent = normalize_excel_key(ws.cell(row_idx, headers["SENT"]).value)
        if sent:
            row_by_sent[sent] = row_idx
    updated, skipped, seen = 0, [], set()
    for xml in sorted(folder.glob("*.xml")):
        if "SENT211" not in xml.stem.upper():
            skipped.append(f"{xml.name}: pominięto, bo to nie SENT211")
            continue
        sent = normalize_excel_key(extract_sent_from_filename(xml.name))
        if not sent or sent in seen or sent not in row_by_sent:
            skipped.append(f"{xml.name}: nie znaleziono SENT w Excelu")
            continue
        seen.add(sent)
        dt = datetime.fromtimestamp(xml.stat().st_ctime)
        cell = ws.cell(row_by_sent[sent], headers["DATA ZAMKNIĘCIA SENT"])
        cell.value = dt
        cell.number_format = "yyyy-mm-dd hh:mm"
        updated += 1
        log(f"OK {xml.name}: {dt:%Y-%m-%d %H:%M}")
    wb.save(excel_path)
    msg = f"Uzupełniono datę zamknięcia SENT dla {updated} wierszy."
    if skipped:
        msg += "\n\nPominięte pliki:\n" + "\n".join(skipped[:30])
    return msg
